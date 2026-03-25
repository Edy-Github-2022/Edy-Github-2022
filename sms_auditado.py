from pathlib import Path
from datetime import datetime
from typing import Callable, Optional
import re
import time

import pandas as pd
from playwright.sync_api import sync_playwright

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle


URL_LOGIN = "https://bbts.bestuse.com.br/#!/login"


# =========================================================
# Helpers de progresso
# =========================================================
def _report_progress(progress_cb: Optional[Callable[[str, int], None]], msg: str, pct: int):
    if progress_cb:
        try:
            progress_cb(msg, pct)
        except Exception:
            pass


# =========================================================
# Helpers Web
# =========================================================
def esperar_tabela(page, seletor_tabela="table.table.table-striped.table-hover", timeout_ms=45000):
    """
    Espera pela presença da tabela (thead e pelo menos 1 linha em tbody).
    Retorna (headers, total_linhas).
    """
    page.wait_for_selector(seletor_tabela, state="attached", timeout=timeout_ms)
    page.wait_for_selector(f"{seletor_tabela} thead th", state="visible", timeout=timeout_ms)

    ths = page.locator(f"{seletor_tabela} thead th")
    th_count = ths.count()
    headers = []

    for i in range(th_count):
        txt = ths.nth(i).inner_text().strip()
        txt = re.sub(r"\s+", " ", txt)
        headers.append(txt)

    elapsed = 0
    poll = 800
    while elapsed < timeout_ms:
        trs = page.locator(f"{seletor_tabela} tbody tr")
        total = trs.count()
        if total and total > 0:
            return headers, total
        page.wait_for_timeout(poll)
        elapsed += poll

    return headers, 0


def abrir_e_capturar_pdu_no_row(row):
    """
    Para a linha atual:
    - Se existir link 'Ver PDU': clica e aguarda o <pre>;
    - Se já estiver 'Ocultar PDU': só lê o <pre>;
    - Retorna o texto do <pre>, ou "".
    """
    try:
        row.scroll_into_view_if_needed()
    except Exception:
        pass

    pre_loc = row.locator("pre.ng-binding, pre")
    if pre_loc.count() > 0:
        try:
            pre_loc.first.wait_for(state="visible", timeout=3000)
            return pre_loc.first.inner_text().strip()
        except Exception:
            pass

    link_pdu = row.get_by_role("link", name=re.compile(r"(Ver|Ocultar)\s+PDU", re.I))
    if link_pdu.count() == 0:
        link_pdu = row.locator("a:has-text('Ver PDU'), a:has-text('Ocultar PDU')")

    if link_pdu.count() > 0:
        try:
            link_text = link_pdu.first.inner_text().strip().lower()
            if "ver pdu" in link_text:
                link_pdu.first.click()
                row.locator("pre").first.wait_for(state="visible", timeout=5000)

            pre_loc = row.locator("pre.ng-binding, pre")
            if pre_loc.count() > 0:
                return pre_loc.first.inner_text().strip()
        except Exception:
            return ""

    return ""


def extrair_tabela(page, seletor_tabela="table.table.table-striped.table-hover"):
    """
    Extrai dados da tabela:
    - lê headers;
    - lê cada linha;
    - se houver coluna PDU, abre e captura o conteúdo;
    - retorna (headers, dados).
    """
    headers, total_linhas = esperar_tabela(page, seletor_tabela=seletor_tabela)
    dados = []

    if total_linhas == 0:
        possivel_vazio = page.locator("text=/sem\\s+registros/i")
        if possivel_vazio.count() > 0:
            print("ℹ️ Consulta retornou 'Sem registros'.")
        else:
            print("⚠️ Nenhuma linha encontrada.")
        return headers, dados

    col_pdu_idx = headers.index("PDU") if "PDU" in headers else -1

    trs = page.locator(f"{seletor_tabela} tbody tr")
    for i in range(total_linhas):
        row = trs.nth(i)
        tds = row.locator("td")
        col_count = tds.count()

        row_vals = []
        for c in range(col_count):
            cell_text = tds.nth(c).inner_text().strip()
            cell_text = re.sub(r"\s+", " ", cell_text)
            row_vals.append(cell_text)

        m = min(len(headers), len(row_vals))
        row_dict = {headers[idx]: row_vals[idx] for idx in range(m)}

        if len(headers) > m:
            for h in headers[m:]:
                row_dict[h] = ""

        if col_pdu_idx >= 0:
            try:
                pdu_json_text = abrir_e_capturar_pdu_no_row(row)
                if pdu_json_text:
                    row_dict["PDU"] = pdu_json_text
            except Exception:
                pass

        dados.append(row_dict)

    return headers, dados


def _fill_first_available(page, selectors, value: str) -> bool:
    for selector in selectors:
        try:
            loc = page.locator(selector)
            if loc.count() > 0:
                loc.first.fill("")
                loc.first.fill(value)
                return True
        except Exception:
            pass
    return False


def _click_first_available(page, selectors) -> bool:
    for selector in selectors:
        try:
            loc = page.locator(selector)
            if loc.count() > 0:
                loc.first.click()
                return True
        except Exception:
            pass
    return False


# =========================================================
# Helpers Excel
# =========================================================
def classificar_transacao(mensagem: str) -> str:
    if not isinstance(mensagem, str) or mensagem.strip() == "":
        return "Sem Transação"

    texto = mensagem.lower()
    tem_valor = bool(re.search(r"r\$\s\d", texto, flags=re.IGNORECASE))

    termos_transacao = [
        "pix", "ted", "doc",
        "pagamento", "pagto",
        "transferência", "transferencia",
        "agendamento", "agendado", "agendada",
        "compra", "compras", "compra aprovada",
        "boleto", "fatura",
        "saque", "depósito", "deposito",
        "cobrança", "cobranca",
        "crédito", "credito",
        "débito", "debito"
    ]

    tem_termo_transacao = any(t in texto for t in termos_transacao)

    if tem_termo_transacao or tem_valor:
        return "Com Transação"

    return "Sem Transação"


def garantir_ordem_colunas(df: pd.DataFrame) -> pd.DataFrame:
    ordem = [
        "Identificador", "Número", "Arquivo", "C.Custo", "Cliente",
        "Data envio do arquivo", "Início do envio", "Fim do envio",
        "Status", "PDU", "Sms do arquivo", "Transação Financeira"
    ]

    for col in ordem:
        if col not in df.columns:
            df[col] = ""

    return df.reindex(columns=ordem)


def aplicar_formatacao_excel(excel_path: str, sheet_name: str = "Dados"):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    last_row = ws.max_row
    last_col = ws.max_column
    last_col_letter = get_column_letter(last_col)
    used_range = f"A1:{last_col_letter}{last_row}"

    if last_row >= 2:
        try:
            table_name = f"TabelaSMS_{int(time.time())}"
            table = Table(displayName=table_name, ref=used_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        except Exception:
            pass

    thin = Side(border_style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[used_range]:
        for cell in row:
            cell.border = border

    for col_letter in ("J", "K"):
        if ws[col_letter + "1"].value is not None:
            for col_cells in ws[f"{col_letter}1:{col_letter}{last_row}"]:
                for cell in col_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    widths = {
        "A": 34,
        "B": 18,
        "C": 60,
        "D": 30,
        "E": 40,
        "F": 22,
        "G": 22,
        "H": 22,
        "I": 16,
        "J": 88,
        "K": 88,
        "L": 22,
    }
    for col_letter, width in widths.items():
        try:
            ws.column_dimensions[col_letter].width = width
        except Exception:
            pass

    if last_row >= 2:
        status_range = f"I2:I{last_row}"

        def add_contains_text_rule(rng: str, texto: str, fill_hex: str):
            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            dxf = DifferentialStyle(fill=fill)
            rule = Rule(type="containsText", operator="containsText", text=texto, dxf=dxf)
            ws.conditional_formatting.add(rng, rule)

        for t in ["NÃO ENTREGUE", "NAO ENTREGUE", "Não entregue", "Nao entregue"]:
            add_contains_text_rule(status_range, t, "FFFF0000")

        for t in ["EXPIRADO", "Expirado"]:
            add_contains_text_rule(status_range, t, "FFFF00")

        for t in ["ENTREGUE", "Entregue"]:
            add_contains_text_rule(status_range, t, "FF00B050")

        for t in ["ENVIADO", "Enviado"]:
            add_contains_text_rule(status_range, t, "FF800080")

    wb.save(excel_path)
    wb.close()


# =========================================================
# Função principal reutilizável
# =========================================================
def executar_consulta(
    numeros: str,
    usuario: str = "",
    senha: str = "",
    url_login: str = URL_LOGIN,
    t_login: int = 15,
    t_token: int = 30,
    outputs_dir: str = "outputs",
    headless: bool = True,
    progress_cb: Optional[Callable[[str, int], None]] = None
):
    """
    Executa a consulta e retorna:
    {
        "headers": [...],
        "dados": [...],
        "excel_path": "...",
        "screenshot_path": "...",
        "total_registros": 0
    }
    """
    numeros_limpos = re.sub(r"[^\d\n]", "", numeros or "").strip()
    if not numeros_limpos:
        raise ValueError("Nenhum número válido foi informado.")

    output_dir = Path(outputs_dir)
    output_dir.mkdir(exist_ok=True)

    primeiro_numero = numeros_limpos.splitlines()[0].strip() if numeros_limpos.splitlines() else "consulta"
    agora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    excel_path = output_dir / f"sms_consulta_celular_{primeiro_numero}_{agora}.xlsx"
    screenshot_path = output_dir / f"evidencia_consulta_celular_{primeiro_numero}_{agora}.png"

    headers = []
    dados = []
    browser = None

    try:
        _report_progress(progress_cb, "Iniciando navegador...", 10)

        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=headless,
                args=[
                    "--no-sandbox",