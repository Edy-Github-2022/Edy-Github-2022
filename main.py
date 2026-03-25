from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import asyncio, os, uuid, zipfile, re, time
from datetime import datetime
from pathlib import Path
import pandas as pd

# Excel
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

app = FastAPI(title="BBTS SMS Consulta")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"]
)

# Pasta de saída
OUTPUTS_DIR = Path("outputs")
OUTPUTS_DIR.mkdir(exist_ok=True)

# Servir o frontend
app.mount("/static", StaticFiles(directory="static"), name="static")

# Página inicial
@app.get("/")
def index():
    return FileResponse("static/index.html")

# ---- API principal (consultar / status / download) ----

jobs = {}

class ConsultaRequest(BaseModel):
    numeros: str
    usuario: str
    senha: str
    url_login: str = "https://bbts.bestuse.com.br/#!/login"
    t_login: int = 15
    t_token: int = 30

@app.post("/consultar")
def iniciar_consulta(req: ConsultaRequest, background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued", "pct": 0, "msg": "Na fila...", "zip": None}
    background_tasks.add_task(
        asyncio.get_event_loop().run_in_executor,
        None, run_consulta, job_id, req
    )
    return {"job_id": job_id}

@app.get("/status/{job_id}")
def status(job_id: str):
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job não encontrado")
    return job

@app.get("/download/{job_id}")
def download(job_id: str):
    job = jobs.get(job_id)
    if not job or job.get("status") != "done":
        raise HTTPException(400, "Job não concluído")

    zip_path = job.get("zip")
    if not zip_path or not Path(zip_path).exists():
        raise HTTPException(404, "Arquivo não encontrado")

    return FileResponse(zip_path, media_type="application/zip", filename=Path(zip_path).name)
